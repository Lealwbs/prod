import openpyxl
import math
import sys
import os
from mip import Model, xsum, minimize, BINARY, INTEGER, CONTINUOUS, OptimizationStatus

def haversine(lat1, lon1, lat2, lon2):
    R = 6371.0
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    a = math.sin(delta_phi / 2.0)**2 + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2.0)**2
    c = 2.0 * math.atan2(math.sqrt(a), math.sqrt(1.0 - a))
    return R * c

def solve_scenario(xlsx_path, m_limit=2, pen_val=100.0):
    wb = openpyxl.load_workbook(xlsx_path)
    
    # Read Parameters
    params = {}
    ws_params = wb['Parametros']
    for row in ws_params.iter_rows(min_row=2, values_only=True):
        if row[0]:
            params[row[0]] = row[1]
    
    cap_caminhao = params.get('capacidade_caminhao_kg', 500.0)
    
    # Read Clients
    clients = []
    ws_clients = wb['Clientes']
    for row in ws_clients.iter_rows(min_row=2, values_only=True):
        if row[0]:
            clients.append({
                'name': row[0],
                'lat': row[1],
                'lon': row[2],
                'demanda': row[3]
            })
            
    # Read Suppliers
    suppliers = []
    ws_suppliers = wb['Fornecedores']
    for row in ws_suppliers.iter_rows(min_row=2, values_only=True):
        if row[0]:
            suppliers.append({
                'name': row[0],
                'lat': row[1],
                'lon': row[2],
                'oferta': row[3]
            })
            
    # Read Candidates
    candidates = []
    ws_candidates = wb['Candidatos']
    for row in ws_candidates.iter_rows(min_row=2, values_only=True):
        if row[0]:
            candidates.append({
                'name': row[0],
                'lat': row[1],
                'lon': row[2],
                'custo_instalacao': row[3]
            })
            
    # Read Roasters
    roasters = []
    ws_roasters = wb['Torrefadoras']
    for row in ws_roasters.iter_rows(min_row=2, values_only=True):
        if row[0]:
            roasters.append({
                'name': row[0],
                'capacidade': row[1],
                'custo_processamento': row[2],
                'tempo_processamento': row[3]
            })
            
    num_suppliers = len(suppliers)
    num_clients = len(clients)
    num_candidates = len(candidates)
    num_roasters = len(roasters)
    
    # Check if total offer >= total demand
    total_offer = sum(s['oferta'] for s in suppliers)
    total_demand = sum(c['demanda'] for c in clients)
    print(f"Total Offer: {total_offer:.4f}")
    print(f"Total Demand: {total_demand:.4f}")
    if total_demand > total_offer:
        print("WARNING: Total demand exceeds total offer. The model will be INFEASIBLE under strict constraints.")
        # We will still try to solve it to see if solver reports infeasible
        
    c_ik = [[0.0 for k in range(num_candidates)] for i in range(num_suppliers)]
    for i in range(num_suppliers):
        for k in range(num_candidates):
            dist = haversine(suppliers[i]['lat'], suppliers[i]['lon'], candidates[k]['lat'], candidates[k]['lon'])
            c_ik[i][k] = dist / cap_caminhao
            
    c_kj = [[0.0 for j in range(num_clients)] for k in range(num_candidates)]
    for k in range(num_candidates):
        for j in range(num_clients):
            dist = haversine(candidates[k]['lat'], candidates[k]['lon'], clients[j]['lat'], clients[j]['lon'])
            c_kj[k][j] = dist / cap_caminhao
            
    model = Model("Coffee_Supply_Chain")
    model.verbose = 0
    
    x = [[model.add_var(name=f"x_{i}_{k}", lb=0.0, var_type=CONTINUOUS) for k in range(num_candidates)] for i in range(num_suppliers)]
    y = [[model.add_var(name=f"y_{k}_{j}", lb=0.0, var_type=CONTINUOUS) for j in range(num_clients)] for k in range(num_candidates)]
    z = [[model.add_var(name=f"z_{k}_{t}", var_type=BINARY) for t in range(num_roasters)] for k in range(num_candidates)]
    w = [model.add_var(name=f"w_{k}", lb=0.0, var_type=INTEGER) for k in range(num_candidates)]
    u = [[[model.add_var(name=f"u_{i}_{k}_{t}", lb=0.0, var_type=CONTINUOUS) for t in range(num_roasters)] for k in range(num_candidates)] for i in range(num_suppliers)]
    e = [model.add_var(name=f"e_{k}", lb=0.0, var_type=CONTINUOUS) for k in range(num_candidates)]
    
    sum_s = sum(s['oferta'] for s in suppliers)
    M_const = sum_s
    
    transport_forn_cd = xsum(c_ik[i][k] * x[i][k] for i in range(num_suppliers) for k in range(num_candidates))
    transport_cd_cli = xsum(c_kj[k][j] * y[k][j] for k in range(num_candidates) for j in range(num_clients))
    cd_installation = xsum(candidates[k]['custo_instalacao'] * w[k] for k in range(num_candidates))
    processing = xsum(roasters[t]['custo_processamento'] * u[i][k][t] for i in range(num_suppliers) for k in range(num_candidates) for t in range(num_roasters))
    inventory_penalty = xsum(pen_val * e[k] for k in range(num_candidates))
    
    model.objective = minimize(transport_forn_cd + transport_cd_cli + cd_installation + processing + inventory_penalty)
    
    # Constraints
    for i in range(num_suppliers):
        model += xsum(x[i][k] for k in range(num_candidates)) <= suppliers[i]['oferta']
    for i in range(num_suppliers):
        model += xsum(x[i][k] for k in range(num_candidates)) == suppliers[i]['oferta']
        
    for j in range(num_clients):
        model += xsum(y[k][j] for k in range(num_candidates)) >= clients[j]['demanda']
    for j in range(num_clients):
        model += xsum(y[k][j] for k in range(num_candidates)) == clients[j]['demanda']
        
    for k in range(num_candidates):
        model += xsum(x[i][k] for i in range(num_suppliers)) == xsum(y[k][j] for j in range(num_clients)) + e[k]
        
    for k in range(num_candidates):
        model += xsum(z[k][t] for t in range(num_roasters)) <= w[k]
        
    for k in range(num_candidates):
        model += xsum(x[i][k] for i in range(num_suppliers)) <= xsum(roasters[t]['capacidade'] * z[k][t] for t in range(num_roasters))
        
    for t in range(num_roasters):
        model += xsum(z[k][t] for k in range(num_candidates)) <= 1
        
    model += xsum(w[k] for k in range(num_candidates)) <= m_limit
    
    for i in range(num_suppliers):
        for k in range(num_candidates):
            model += xsum(u[i][k][t] for t in range(num_roasters)) == x[i][k]
            
    for i in range(num_suppliers):
        for k in range(num_candidates):
            for t in range(num_roasters):
                model += u[i][k][t] <= x[i][k]
                
    for i in range(num_suppliers):
        for k in range(num_candidates):
            for t in range(num_roasters):
                model += u[i][k][t] <= M_const * z[k][t]
                
    status = model.optimize()
    print("Optimization Status:", status)
    if status == OptimizationStatus.OPTIMAL or status == OptimizationStatus.FEASIBLE:
        obj_val = model.objective_value
        val_transport_forn_cd = sum(c_ik[i][k] * x[i][k].x for i in range(num_suppliers) for k in range(num_candidates))
        val_transport_cd_cli = sum(c_kj[k][j] * y[k][j].x for k in range(num_candidates) for j in range(num_clients))
        val_cd_installation = sum(candidates[k]['custo_instalacao'] * w[k].x for k in range(num_candidates))
        val_processing = sum(roasters[t]['custo_processamento'] * u[i][k][t].x for i in range(num_suppliers) for k in range(num_candidates) for t in range(num_roasters))
        val_inventory_penalty = sum(pen_val * e[k].x for k in range(num_candidates))
        
        print(f"Custo Total: {obj_val:.2f}")
        print(f"  Transporte Fornecedor -> CD: {val_transport_forn_cd:.2f}")
        print(f"  Transporte CD -> Cliente: {val_transport_cd_cli:.2f}")
        print(f"  Instalacao CDs: {val_cd_installation:.2f}")
        print(f"  Processamento Torrefadoras: {val_processing:.2f}")
        print(f"  Penalidade Estoque: {val_inventory_penalty:.2f}")
        
        print("Open CDs:")
        for k in range(num_candidates):
            if w[k].x > 0.01:
                active_roasters = [roasters[t]['name'] for t in range(num_roasters) if z[k][t].x > 0.5]
                print(f"  {candidates[k]['name']}: w_k = {w[k].x}, roasters = {active_roasters}, Entrada = {sum(x[i][k].x for i in range(num_suppliers)):.2f}, Saida = {sum(y[k][j].x for j in range(num_clients)):.2f}, Estoque = {e[k].x:.2f}")
                
        # Print non-zero flows
        print("\nFlows Fornecedor -> CD:")
        for i in range(num_suppliers):
            for k in range(num_candidates):
                if x[i][k].x > 0.01:
                    print(f"  {suppliers[i]['name']} -> {candidates[k]['name']}: {x[i][k].x:.2f}")
                    
        print("\nFlows CD -> Cliente:")
        for k in range(num_candidates):
            for j in range(num_clients):
                if y[k][j].x > 0.01:
                    print(f"  {candidates[k]['name']} -> {clients[j]['name']}: {y[k][j].x:.2f}")
    else:
        print("No solution.")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python solve_single.py <xlsx_path> [m_limit] [pen_val]")
        sys.exit(1)
    path = sys.argv[1]
    m_lim = int(sys.argv[2]) if len(sys.argv) > 2 else 2
    pen = float(sys.argv[3]) if len(sys.argv) > 3 else 100.0
    print(f"Solving {path} with m={m_lim}, pen={pen}...")
    solve_scenario(path, m_lim, pen)
