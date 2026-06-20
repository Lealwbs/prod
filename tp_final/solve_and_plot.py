import openpyxl
import math
import os
import matplotlib.pyplot as plt
from mip import Model, xsum, minimize, BINARY, INTEGER, CONTINUOUS, OptimizationStatus

def haversine(lat1, lon1, lat2, lon2):
    R = 6371.0
    phi1 = math.radians(lat1)
    phi2 = math.radians(lat2)
    delta_phi = math.radians(lat2 - lat1)
    delta_lambda = math.radians(lon2 - lon1)
    a = math.sin(delta_phi / 2.0)**2 + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2.0)**2
    c = 2.0 * math.atan2(math.sqrt(a), math.sqrt(1.0 - a))
    return R * c

def solve_and_plot(xlsx_path, output_png_path, m_limit=2, pen_val=100.0, relax_supply=False, title=""):
    wb = openpyxl.load_workbook(xlsx_path)
    
    # Read Parameters
    params = {}
    ws_params = wb['Parametros']
    for row in ws_params.iter_rows(min_row=2, values_only=True):
        if row[0]:
            params[row[0]] = row[1]
    cap_caminhao = params.get('capacidade_caminhao_kg', 500.0)
    
    # Read Clients
    clients = []
    ws_clients = wb['Clientes']
    for row in ws_clients.iter_rows(min_row=2, values_only=True):
        if row[0]:
            clients.append({'name': row[0], 'lat': row[1], 'lon': row[2], 'demanda': row[3]})
            
    # Read Suppliers
    suppliers = []
    ws_suppliers = wb['Fornecedores']
    for row in ws_suppliers.iter_rows(min_row=2, values_only=True):
        if row[0]:
            suppliers.append({'name': row[0], 'lat': row[1], 'lon': row[2], 'oferta': row[3]})
            
    # Read Candidates
    candidates = []
    ws_candidates = wb['Candidatos']
    for row in ws_candidates.iter_rows(min_row=2, values_only=True):
        if row[0]:
            candidates.append({'name': row[0], 'lat': row[1], 'lon': row[2], 'custo_instalacao': row[3]})
            
    # Read Roasters
    roasters = []
    ws_roasters = wb['Torrefadoras']
    for row in ws_roasters.iter_rows(min_row=2, values_only=True):
        if row[0]:
            roasters.append({'name': row[0], 'capacidade': row[1], 'custo_processamento': row[2], 'tempo_processamento': row[3]})
            
    num_suppliers = len(suppliers)
    num_clients = len(clients)
    num_candidates = len(candidates)
    num_roasters = len(roasters)
    
    c_ik = [[0.0 for k in range(num_candidates)] for i in range(num_suppliers)]
    for i in range(num_suppliers):
        for k in range(num_candidates):
            dist = haversine(suppliers[i]['lat'], suppliers[i]['lon'], candidates[k]['lat'], candidates[k]['lon'])
            c_ik[i][k] = dist / cap_caminhao
            
    c_kj = [[0.0 for j in range(num_clients)] for k in range(num_candidates)]
    for k in range(num_candidates):
        for j in range(num_clients):
            dist = haversine(candidates[k]['lat'], candidates[k]['lon'], clients[j]['lat'], clients[j]['lon'])
            c_kj[k][j] = dist / cap_caminhao
            
    model = Model("Coffee_Supply_Chain")
    model.verbose = 0
    
    x = [[model.add_var(name=f"x_{i}_{k}", lb=0.0, var_type=CONTINUOUS) for k in range(num_candidates)] for i in range(num_suppliers)]
    y = [[model.add_var(name=f"y_{k}_{j}", lb=0.0, var_type=CONTINUOUS) for j in range(num_clients)] for k in range(num_candidates)]
    z = [[model.add_var(name=f"z_{k}_{t}", var_type=BINARY) for t in range(num_roasters)] for k in range(num_candidates)]
    w = [model.add_var(name=f"w_{k}", lb=0.0, var_type=INTEGER) for k in range(num_candidates)]
    u = [[[model.add_var(name=f"u_{i}_{k}_{t}", lb=0.0, var_type=CONTINUOUS) for t in range(num_roasters)] for k in range(num_candidates)] for i in range(num_suppliers)]
    e = [model.add_var(name=f"e_{k}", lb=0.0, var_type=CONTINUOUS) for k in range(num_candidates)]
    
    sum_s = sum(s['oferta'] for s in suppliers)
    M_const = sum_s
    
    transport_forn_cd = xsum(c_ik[i][k] * x[i][k] for i in range(num_suppliers) for k in range(num_candidates))
    transport_cd_cli = xsum(c_kj[k][j] * y[k][j] for k in range(num_candidates) for j in range(num_clients))
    cd_installation = xsum(candidates[k]['custo_instalacao'] * w[k] for k in range(num_candidates))
    processing = xsum(roasters[t]['custo_processamento'] * u[i][k][t] for i in range(num_suppliers) for k in range(num_candidates) for t in range(num_roasters))
    inventory_penalty = xsum(pen_val * e[k] for k in range(num_candidates))
    
    model.objective = minimize(transport_forn_cd + transport_cd_cli + cd_installation + processing + inventory_penalty)
    
    # Constraints
    for i in range(num_suppliers):
        model += xsum(x[i][k] for k in range(num_candidates)) <= suppliers[i]['oferta']
    if not relax_supply:
        for i in range(num_suppliers):
            model += xsum(x[i][k] for k in range(num_candidates)) == suppliers[i]['oferta']
        
    for j in range(num_clients):
        model += xsum(y[k][j] for k in range(num_candidates)) >= clients[j]['demanda']
    for j in range(num_clients):
        model += xsum(y[k][j] for k in range(num_candidates)) == clients[j]['demanda']
        
    for k in range(num_candidates):
        model += xsum(x[i][k] for i in range(num_suppliers)) == xsum(y[k][j] for j in range(num_clients)) + e[k]
        
    for k in range(num_candidates):
        model += xsum(z[k][t] for t in range(num_roasters)) <= w[k]
        
    for k in range(num_candidates):
        model += xsum(x[i][k] for i in range(num_suppliers)) <= xsum(roasters[t]['capacidade'] * z[k][t] for t in range(num_roasters))
        
    for t in range(num_roasters):
        model += xsum(z[k][t] for k in range(num_candidates)) <= 1
        
    model += xsum(w[k] for k in range(num_candidates)) <= m_limit
    
    for i in range(num_suppliers):
        for k in range(num_candidates):
            model += xsum(u[i][k][t] for t in range(num_roasters)) == x[i][k]
            
    for i in range(num_suppliers):
        for k in range(num_candidates):
            for t in range(num_roasters):
                model += u[i][k][t] <= x[i][k]
                
    for i in range(num_suppliers):
        for k in range(num_candidates):
            for t in range(num_roasters):
                model += u[i][k][t] <= M_const * z[k][t]
                
    status = model.optimize()
    if status != OptimizationStatus.OPTIMAL and status != OptimizationStatus.FEASIBLE:
        print("Model could not be solved.")
        return
        
    # Plotting
    plt.figure(figsize=(10, 8))
    
    # Extract coordinates
    supp_lons = [s['lon'] for s in suppliers]
    supp_lats = [s['lat'] for s in suppliers]
    cli_lons = [c['lon'] for c in clients]
    cli_lats = [c['lat'] for c in clients]
    
    # Plot all clients
    plt.scatter(cli_lons, cli_lats, color='blue', marker='o', s=50, label='Clientes', alpha=0.6)
    for c in clients:
        plt.text(c['lon'] + 0.05, c['lat'] + 0.05, c['name'], fontsize=8, color='blue')
        
    # Plot all suppliers
    plt.scatter(supp_lons, supp_lats, color='green', marker='^', s=80, label='Fornecedores', alpha=0.8)
    for s in suppliers:
        plt.text(s['lon'] - 0.15, s['lat'] - 0.15, s['name'], fontsize=8, color='green')
        
    # Plot active CDs
    active_cd_indices = []
    for k in range(num_candidates):
        if w[k].x > 0.01:
            active_cd_indices.append(k)
            plt.scatter(candidates[k]['lon'], candidates[k]['lat'], color='red', marker='s', s=150, edgecolors='black', linewidths=1.5, zorder=5)
            plt.text(candidates[k]['lon'] + 0.08, candidates[k]['lat'] - 0.08, f"CD {candidates[k]['name']} (w={int(w[k].x)})", fontsize=10, fontweight='bold', color='red')
            
    # Draw flows
    # Fornecedor -> CD (dashed lines)
    for i in range(num_suppliers):
        for k in range(num_candidates):
            if x[i][k].x > 0.01:
                plt.plot([suppliers[i]['lon'], candidates[k]['lon']], [suppliers[i]['lat'], candidates[k]['lat']], color='green', linestyle='--', alpha=0.5, linewidth=1.5)
                # print travel size
                mid_lon = (suppliers[i]['lon'] + candidates[k]['lon']) / 2.0
                mid_lat = (suppliers[i]['lat'] + candidates[k]['lat']) / 2.0
                plt.text(mid_lon, mid_lat, f"{x[i][k].x:.0f} kg", fontsize=7, color='green', bbox=dict(facecolor='white', alpha=0.7, boxstyle='round,pad=0.2'))

    # CD -> Cliente (solid lines)
    for k in range(num_candidates):
        for j in range(num_clients):
            if y[k][j].x > 0.01:
                plt.plot([candidates[k]['lon'], clients[j]['lon']], [candidates[k]['lat'], clients[j]['lat']], color='red', linestyle='-', alpha=0.6, linewidth=1.5)
                mid_lon = (candidates[k]['lon'] + clients[j]['lon']) / 2.0
                mid_lat = (candidates[k]['lat'] + clients[j]['lat']) / 2.0
                plt.text(mid_lon, mid_lat, f"{y[k][j].x:.0f} kg", fontsize=7, color='red', bbox=dict(facecolor='white', alpha=0.7, boxstyle='round,pad=0.2'))
                
    plt.title(title, fontsize=12, fontweight='bold')
    plt.xlabel("Longitude")
    plt.ylabel("Latitude")
    plt.grid(True, linestyle=':', alpha=0.6)
    
    # Custom legend
    plt.scatter([], [], color='red', marker='s', s=100, label='CD Instalado')
    plt.plot([], [], color='green', linestyle='--', label='Fluxo Forn -> CD')
    plt.plot([], [], color='red', linestyle='-', label='Fluxo CD -> Cli')
    plt.legend(loc='best')
    
    plt.tight_layout()
    plt.savefig(output_png_path, dpi=150)
    plt.close()
    print(f"Map saved to {output_png_path}")

if __name__ == "__main__":
    xlsx = "tp_final/cenarios/cenario_2_aumento_oferta.xlsx"
    os.makedirs("tp_final/imgs", exist_ok=True)
    
    print("Solving and plotting Scenario 2 (Strict)...")
    solve_and_plot(xlsx, "tp_final/imgs/mapa_cenario_2.png", m_limit=2, relax_supply=False, title="Rotas Geradas no Cenário 2 (Aumento da Oferta - Rígido)")
    
    print("Solving and plotting Scenario 2 Extra (Relaxed)...")
    solve_scenario_extra = solve_and_plot(xlsx, "tp_final/imgs/mapa_cenario_2_extra.png", m_limit=2, relax_supply=True, title="Rotas Geradas no Cenário 2 Extra (Aumento da Oferta - Flexível)")
